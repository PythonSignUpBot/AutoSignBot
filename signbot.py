import openpyxl
import random
 
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("signupinfo.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        if col[row].value == None:
            if col[row].column == 2:
                day = random.randint(1, 28)
                month = random.randint(1, 12)
                year = random.randint(1958, 1994)
                birthday = str(day) + '/' + str(month) + '/' + str(year)
                dataframe1.cell(row, 2).value = 3
                print(col[row].value)

dataframe.save("signupinfo.xlsx")