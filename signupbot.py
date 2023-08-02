import openpyxl
import random
import requests
import json
 
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("signupinfo.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
colName = ['A', 'B', 'C', 'D', 'E', 'F']
prevPhoneNumber = ["+075 ", "+074 "]
url = 'https://app.anonaddy.com/api/v1/aliases'
payload = {
    "domain": "anonaddy.me",
    "description": "For example.com",
    "format": "random_characters",
    "local_part": "hello"
}
headers = {
  'Content-Type': 'application/json',
  'X-Requested-With': 'XMLHttpRequest',
  'Authorization': 'Bearer rYgW9f1sobflAsWGsZlaKZhs6OpsvvAkwoQUFyry'
}
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        if col[row].value == None:
            if col[row].column == 2:
              day = random.randint(1, 28)
              month = random.randint(1, 12)
              year = random.randint(1958, 1994)
              birthday = str(day) + '/' + str(month) + '/' + str(year)
              cellName = colName[col[row].column - 1]+str(row+1)
              dataframe1[cellName].value = birthday
            elif col[row].column == 6:
              prevNum = random.randint(0, 1)
              backPhoneNumber = random.randint(12345678, 98765432)
              phoneNumber = prevPhoneNumber[prevNum] + str(backPhoneNumber)
              print(phoneNumber)
              cellName = 'F' + str(row+1)
              dataframe1[cellName].value = phoneNumber
            elif col[row].column == 5:
              response = requests.post(url, headers=headers, json=payload)
              data = response.json()
              cellName = 'E' + str(row+1)
              dataframe1[cellName].value = data['data']['email']

dataframe.save("signupinfo.xlsx")