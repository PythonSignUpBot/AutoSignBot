import openpyxl
import random
import requests
import json
 
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("signupinfo.xlsx")
# Define variable to read sheet
dataframe1 = dataframe.active
vName = ''
vBirthday = ''
vPostCode = ''
vAddress = ''
vMail = ''
vPhone = ''
for row in range(1, dataframe1.max_row):
  for col in dataframe1.iter_cols(1, dataframe1.max_column):
    if col[row].column == 1:
      vName = col[row].value
    elif col[row].column == 2:
      vBirthday = col[row].value
    elif col[row].column == 3:
      vPostCode = col[row].value
    elif col[row].column == 4:
      vAddress = col[row].value
    elif col[row].column == 5:
      vMail = col[row].value
    elif col[row].column == 6:
      vPhone = col[row].value
  print(vName, vBirthday)